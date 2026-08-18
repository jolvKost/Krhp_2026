"""Escanea formulas de un Excel (nunca datos) para mapear un proceso externo.

Corre 100% local. Genera un .txt con encabezados de fila 1 y texto de
formulas unicamente -- las celdas con valores planos nunca se leen.
"""

import argparse
import datetime
from pathlib import Path

from openpyxl import load_workbook

ADVERTENCIA = """=== ADVERTENCIA ===
Este reporte contiene UNICAMENTE texto de formulas, nunca valores de celda.
Una formula puede incluir un literal embebido, ej: =IF(A2="Juan Perez",...)
Este script NO detecta ni redacta datos personales automaticamente
(deliberado: un detector heuristico de PII no es confiable).
Revisa este archivo visualmente (~1 min) antes de compartir cualquier fragmento.
"""


def _sheet_headers(ws) -> list[str]:
    fila1 = next(ws.iter_rows(min_row=1, max_row=1), [])
    return [str(c.value) if c.value is not None else "" for c in fila1]


def _formula_cells(ws) -> list[tuple[str, str]]:
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if getattr(cell, "data_type", None) == "f":
                formulas.append((cell.coordinate, cell.value))
    return formulas


def scan_workbook(path) -> str:
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        sheets = [(ws.title, _sheet_headers(ws), _formula_cells(ws)) for ws in wb.worksheets]
    finally:
        wb.close()
    return _format_report(path, sheets)


def _format_report(path, sheets) -> str:
    lineas = [
        f"Reporte de formulas - scan_excel_logic.py",
        f"Archivo: {path}",
        f"Generado: {datetime.datetime.now():%Y-%m-%d %H:%M:%S}",
        "",
        ADVERTENCIA,
    ]
    for nombre, encabezados, formulas in sheets:
        lineas.append(f"=== Hoja: {nombre} ===")
        lineas.append(f"Encabezados (fila 1): {' | '.join(encabezados)}")
        lineas.append(f"Formulas encontradas: {len(formulas)}")
        for coord, formula in formulas:
            lineas.append(f"  {coord} -> {formula}")
        lineas.append("")
    return "\n".join(lineas)


def _selfcheck() -> None:
    import tempfile

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws["A1"] = "Nombre"
    ws["B1"] = "Total"
    ws["A2"] = "Maria Lopez"
    ws["B2"] = "=SUM(1,2)"

    with tempfile.TemporaryDirectory() as d:
        ruta = Path(d) / "test.xlsx"
        wb.save(ruta)
        reporte = scan_workbook(ruta)

    assert "Nombre" in reporte and "Total" in reporte
    assert "B2 -> =SUM(1,2)" in reporte
    assert "Maria Lopez" not in reporte
    print("selfcheck OK")


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", nargs="?", help="Ruta al archivo .xlsx a escanear")
    parser.add_argument("-o", "--output", default="scan_report.txt", help="Archivo de salida")
    parser.add_argument("--selfcheck", action="store_true", help="Corre el autocheck y sale")
    args = parser.parse_args(argv)

    if args.selfcheck:
        _selfcheck()
        return 0

    ruta = args.workbook or input("Ruta al archivo .xlsx: ").strip()
    reporte = scan_workbook(ruta)
    Path(args.output).write_text(reporte, encoding="utf-8")
    print(f"Reporte guardado en {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
